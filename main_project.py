# -*- coding: utf-8 -*-
"""
Created on Thu Oct  1 13:49:50 2020

@author: Connor
"""

from tkinter import *
import serial
import time
import pandas as pd
from scipy.signal import argrelextrema
import numpy as np

from datetime import datetime
import tkinter.font as tkFont
import matplotlib.pyplot as plt
import matplotlib.animation as animation
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from openpyxl import load_workbook
from openpyxl import Workbook

import analyze_data

workbook_file_name = 'force_production_data_deadlift.xlsx'

def open_files():
    #filename = 'analyze_data.py'
    #with open(filename, "rb") as source_file:
    #    code = compile(source_file.read(), filename, "exec")
    #exec(code)
    analyze_data.SelectionInput()

class MainGUI:

    def __init__(self):
        self.root = Tk()
        self.root.geometry("900x800")
        self.root.title("Force Produciton Tracker")
        self.root.configure(bg='black')
        
        _gray = '#595a61'
        message_font_style = tkFont.Font(size=30)
        
        ### Variables
        self.run_animation = False
        self.count = 0
        self.continuous = True
        self.mass_entered = 0
        self.max_acceleration = 0
        self.max_force = 0
        self.lift_style = ""
        
        ### Frames
        self.title_frame = Frame(self.root)
        self.title_frame.grid(column=0, row=0)
        self.title_frame.configure(bg='black')
        
        self.middle_frame = Frame(self.root)
        self.middle_frame.grid(column=0, row=1)
        self.middle_frame.configure(bg='black')
        
        self.graph_frame = Frame(self.middle_frame)
        self.graph_frame.grid(column=0, row=0)
        self.graph_frame.configure(bg='black')
        
        self.button_frame = Frame(self.graph_frame)
        self.button_frame.grid(column=0, row=1)
        self.button_frame.configure(bg='black')
        
        self.data_display_frame = Frame(self.middle_frame)
        self.data_display_frame.grid(column=1, row=0)
        self.data_display_frame.configure(bg=_gray)
        
        self.data_entry_frame = Frame(self.data_display_frame)
        self.data_entry_frame.grid(column=0, row=1, padx=10, pady=10)
        self.data_entry_frame.configure(bg=_gray)
        
        self.out_display_frame = Frame(self.data_display_frame)
        self.out_display_frame.grid(column=0, row=2)
        self.out_display_frame.configure(bg=_gray)
        
        self.out_display_labels_frame = Frame(self.out_display_frame)
        self.out_display_labels_frame.grid(column=0, row=0)
        self.out_display_labels_frame.configure(bg=_gray)
        
        self.out_display_data_frame = Frame(self.out_display_frame)
        self.out_display_data_frame.grid(column=1, row=0)
        self.out_display_data_frame.configure(bg=_gray)
        
        
        ### Messages
        self.title_message = Message(self.title_frame, font=message_font_style)
        self.title_message.grid(column=0, row=0)
        self.title_message.configure(text="Force Production Tracker") 
        self.title_message.configure(bg='black', fg='orange')
        self.title_message.configure(width=900)  
        
        self.mass_label = Label(self.out_display_labels_frame, text='Mass (lbs):')
        self.mass_label.grid(column=0, row=0)
        self.mass_label.configure(width=16, anchor='w', bg=_gray, fg='orange')
        
        self.max_acceleration_label = Label(self.out_display_labels_frame, text='Max Acceleration (g):')
        self.max_acceleration_label.grid(column=0, row=1)
        self.max_acceleration_label.configure(width=16, anchor='w', bg=_gray, fg='orange')
        
        self.max_force_label = Label(self.out_display_labels_frame, text='Max Force (N):')
        self.max_force_label.grid(column=0, row=2)
        self.max_force_label.configure(width=16, anchor='w', bg=_gray, fg='orange')
        
        self.lift_style_label = Label(self.out_display_labels_frame, text='Lift Style:')
        self.lift_style_label.grid(column=0, row=3)
        self.lift_style_label.configure(width=16, anchor='w', bg=_gray, fg='orange')
        
        self.mass_data_label = Label(self.out_display_data_frame, text=str(self.mass_entered))
        self.mass_data_label.grid(column=0, row=0)
        self.mass_data_label.configure(width=6, anchor='e', bg=_gray, fg='orange')
        
        self.max_acceleration_data_label = Label(self.out_display_data_frame, text=str(self.max_acceleration))
        self.max_acceleration_data_label.grid(column=0, row=1)
        self.max_acceleration_data_label.configure(width=6, anchor='e', bg=_gray, fg='orange')
        
        self.max_force_data_label = Label(self.out_display_data_frame, text=str(self.max_force))
        self.max_force_data_label.grid(column=0, row=2)
        self.max_force_data_label.configure(width=6, anchor='e', bg=_gray, fg='orange')
        
        self.selected_style = StringVar(self.out_display_data_frame)
        self.list_of_lift_styles = ["Normal", "Slow", "Speed", "Chain", "Negative", "Reverse", "Football"]
        self.lift_style_spinbox = OptionMenu(self.out_display_data_frame, self.selected_style, *self.list_of_lift_styles, command=self.select_style)
        self.lift_style_spinbox.grid(column=0, row=3)
        self.lift_style_spinbox.configure(bg=_gray, fg='orange', width=8)
        
        ### Buttons
        self.start_button = Button(self.button_frame)
        self.start_button.grid(column=0, row=0, padx=2)
        self.start_button.configure(text="Start")
        self.start_button.configure(fg='orange', bg=_gray, relief='raised')
        self.start_button.configure(command=self.click_start)
        
        self.stop_button = Button(self.button_frame)
        self.stop_button.grid(column=1, row=0, padx=2)
        self.stop_button.configure(text="Stop")
        self.stop_button.configure(fg='orange', bg=_gray, relief='raised')
        self.stop_button.configure(command=self.click_stop)
        
        self.continuous_button = Button(self.button_frame)
        self.continuous_button.grid(column=2, row=0, padx=2)
        self.continuous_button.configure(text="Toggle Continuous")
        self.continuous_button.configure(fg='orange', bg=_gray, relief='raised')
        self.continuous_button.configure(command=self.click_continuous)
        
        self.reset_button = Button(self.button_frame)
        self.reset_button.grid(column=3, row=0, padx=2)
        self.reset_button.configure(text="Reset")
        self.reset_button.configure(fg='orange', bg=_gray, relief='raised')
        self.reset_button.configure(command=self.click_reset)
    
        self.save_button = Button(self.button_frame)
        self.save_button.grid(column=4, row=0, padx=2)
        self.save_button.configure(text="Save")
        self.save_button.configure(fg='orange', bg=_gray, relief='raised')
        self.save_button.configure(command=self.click_save)
        
        self.open_button = Button(self.button_frame)
        self.open_button.grid(column=5, row=0, padx=2)
        self.open_button.configure(text='Open')
        self.open_button.configure(fg='orange', bg=_gray, relief='raised')
        self.open_button.configure(command=self.click_open)
        
        self.quit_button = Button(self.button_frame)
        self.quit_button.grid(column=6, row=0, padx=2)
        self.quit_button.configure(text="Quit")
        self.quit_button.configure(fg='orange', bg=_gray, relief='raised')
        self.quit_button.configure(command=self.click_quit)
    
    
        ### Mass Entry Box 
        self.mass_entry_label = Label(self.data_display_frame)
        self.mass_entry_label.grid(column=0, row=0, padx=5, pady=5)
        self.mass_entry_label.configure(text='Enter Mass Used', fg='orange', bg=_gray)
        
        self.mass_entry = Entry(self.data_entry_frame)
        self.mass_entry.grid(column=0, row=0, padx=5, pady=5)
        self.mass_entry.configure(width=8)
        
        self.mass_entry.bind('<Return>', self.click_enter)
        
        self.mass_entry_button = Button(self.data_entry_frame)
        self.mass_entry_button.grid(column=1, row=0, padx=5, pady=5)
        self.mass_entry_button.configure(text='Enter')
        self.mass_entry_button.configure(fg='orange', bg=_gray, relief='raised')
        self.mass_entry_button.configure(command=self.click_enter)
        
        
        ##Graph and Data setup
        self.data = DataUpdate()
        self.graph_info = GraphSetup()

        self.plot_canvas = FigureCanvasTkAgg(self.graph_info.fig, self.graph_frame)
        self.plot_canvas.get_tk_widget().grid(column=0, row=0)
        self.plot_canvas.get_tk_widget().configure(bg='black')
        
        ani = animation.FuncAnimation(self.graph_info.fig, self.data_update, interval=25)


        self.root.mainloop()
        
    
    def data_update(self, i):
        if (self.run_animation == True):
            xdata, ydata, xcdata, ycdata = self.data.add_data()
            self.graph_info.set_data(xdata, ydata, xcdata, ycdata, self.continuous)
            self.calculate_data(xdata, ydata)
            
    def click_start(self):
        self.run_animation = True
    def click_stop(self):
        self.run_animation = False
    def click_continuous(self):
        self.continuous = not self.continuous
    def click_reset(self):
        self.data.reset_data()
        self.graph_info.reset_graph()
        self.graph_info.set_data([],[],[],[],self.continuous)
    def click_save(self):
        xdata, ydata, a, b = self.data.add_data()
        new_save = Save_Data_GUI(xdata, ydata, self.mass_entered, self.lift_style)
    def click_open(self):
        open_files()
    def click_quit(self):
        self.data._quit()
        self.root.quit()
        self.root.destroy()      
    def select_style(self, selection):
        self.lift_style = selection
    def click_enter(self, event=None):
        try:
            self.mass_entered = int(self.mass_entry.get())
        except:
            pass
        
        if isinstance(self.mass_entered, int):
            self.mass_data_label.configure(text=self.mass_entered)
            #self.calculate_data()
    def calculate_data(self, x, y):
        n = 10
        data = pd.DataFrame(list(zip(x, y)), columns=['x', 'y'])
        data['max'] = data.iloc[argrelextrema(data['y'].values, np.greater, order=n)[0]]['y']
        peak_accel = data['max'].max() - data['y'][0:200].mean() #in g, 1 g = 9.80665 N/kg
        peak_force = (peak_accel * 9.80665) * (self.mass_entered / 2.205)
        
        #print("peak accel: ", peak_accel)
        #print("peak force: ", peak_force)
        
        self.max_acceleration_data_label.configure(text=round(peak_accel, 2))
        self.max_force_data_label.configure(text=round(peak_force, 2))
        
class GraphSetup():
    def __init__(self):
        self.fig, self.ax = plt.subplots(figsize=(9,9))
        self.reset_graph()
        
        #need to include x and y
        self.line, = self.ax.plot([], [], color='orange', marker='o')
      
        plt.style.use('dark_background')
        
    def set_data(self, x, y, xc, yc, continuous):
        if continuous:
            self.line.set_data(x, y)
        else:
            self.line.set_data(xc, yc)
            
        if (len(x) > 0 and continuous):
            self.ax.set_xlim(0, x[-1])
        elif not continuous:
            if (len(xc) > 1):
                self.ax.set_xlim(xc[0], xc[-1])
        
    def reset_graph(self):
        self.ax.set_ylim(-2,2)
        self.ax.set_xlim(0,.5)   

class DataUpdate():
    def __init__(self):
        self.baud_rate = 115200
        self.ser_port = 'COM3'
        self.ser = serial.Serial(self.ser_port, self.baud_rate)
        self.reset_data()
                
    def reset_data(self):
        self.current_x_data = []
        self.current_y_data = []
        self.tot_x_data = []
        self.tot_y_data = []
        self.count = 0
        self.append_count = 0
        
        self.connect_serial()
        
    def connect_serial(self): 
        self.count = 0
        self.ser.close()
        self.ser.open()            
        time.sleep(1)
    
    def get_data(self):
        return self.tot_x_data, self.tot_y_data

    def add_data(self):
        ## TODO: add logic to delete beignning of current x and y data
        self.ser.reset_input_buffer()
        data_in = self.ser.readline().decode('utf-8').split()
        number_current_data_points = 100
        #print(data_in)
        
        try:
            data_in = float(data_in[2])
        except:
            pass
        
        if isinstance(data_in, float):
            self.count += 1
            self.append_count = self.count / 20
            self.current_x_data.append(self.append_count)
            self.current_y_data.append(data_in)
            self.tot_x_data.append(self.append_count)
            self.tot_y_data.append(data_in)
            
            if (len(self.current_x_data) > number_current_data_points):
                self.current_x_data.pop(0)
                self.current_y_data.pop(0)
                #print(len(self.current_x_data))
                
            
            
            
        return self.tot_x_data, self.tot_y_data, self.current_x_data, self.current_y_data
            
    def _quit(self):
        self.ser.close()

class Save_Data_GUI:
    def __init__(self, TOT_X_DATA, TOT_Y_DATA, MASS, STYLE):
        self.root = Tk()
        self.root.geometry("350x200")
        self.root.title("Save Force Production Data")
        self.root.configure(bg='black')
        
        #print("in save class")
        
        self.tot_x_data = TOT_X_DATA
        self.tot_y_data = TOT_Y_DATA
        self.mass = MASS
        self.lift_style = STYLE
        
        _gray = '#595a61'
        title_message_font_style = tkFont.Font(size=20)
        date = datetime.today().strftime('%Y-%m-%d')
        
        ### Frames
        self.title_frame = Frame(self.root)
        self.title_frame.grid(column=0, row=0)
        self.title_frame.configure(bg='black')
        
        self.container_frame = Frame(self.root)
        self.container_frame.grid(column=0, row=1)
        self.container_frame.configure(bg='black')
        
        self.button_frame = Frame(self.root)
        self.button_frame.grid(column=0, row=2)
        self.button_frame.configure(bg='black', pady=10)
        
        ### Messages
        self.title_message = Message(self.title_frame, font=title_message_font_style)
        self.title_message.grid(column=0, row=0)
        self.title_message.configure(text="Save Data")
        self.title_message.configure(bg='black', fg='orange', width=300, padx=115)
        
        label_width = 7
        
        self.name_label = Label(self.container_frame)
        self.name_label.grid(column=0, row=0)
        self.name_label.configure(text='Name: ', bg='black', fg='orange', width=label_width, anchor='w')
        
        self.mass_label = Label(self.container_frame)
        self.mass_label.grid(column=0, row=1)
        self.mass_label.configure(text='Mass: ', bg='black', fg='orange', width=label_width, anchor='w')
        
        self.reps_label = Label(self.container_frame)
        self.reps_label.grid(column=0, row=2)
        self.reps_label.configure(text='Reps: ', bg='black', fg='orange', width=label_width, anchor='w')
        
        self.lift_style_label = Label(self.container_frame)
        self.lift_style_label.grid(column=0, row=3)
        self.lift_style_label.configure(text='Lift Style:', bg='black', fg='orange', width=label_width, anchor='w')
        
        self.date_label = Label(self.container_frame)
        self.date_label.grid(column=0, row=4)
        self.date_label.configure(text='Date: ', bg='black', fg='orange', width=label_width, anchor='w')
        
        self.save_file_label = Label(self.container_frame)
        self.save_file_label.grid(column=0, row=5)
        self.save_file_label.configure(text='File: ', bg='black', fg='orange', width=label_width, anchor='w')
        
        
        
        ### Entries
        self.name_entry = Entry(self.container_frame)
        self.name_entry.grid(column=1, row=0)
        self.name_entry.bind('<Return>', self.click_save)

        self.mass_entry = Entry(self.container_frame)
        self.mass_entry.grid(column=1, row=1)
        self.mass_entry.insert(END, self.mass)
        self.mass_entry.bind('<Return>', self.click_save)
        
        self.reps_entry = Entry(self.container_frame)
        self.reps_entry.grid(column=1, row=2)
        self.reps_entry.bind('<Return>', self.click_save)

        self.style_entry = Entry(self.container_frame)
        self.style_entry.grid(column=1, row=3)
        self.style_entry.insert(END, self.lift_style)
        self.style_entry.bind('<Return>', self.click_save)
        
        self.date_entry = Entry(self.container_frame)
        self.date_entry.grid(column=1, row=4)
        self.date_entry.insert(END, date)
        self.date_entry.bind('<Return>', self.click_save)
        
        self.save_file_entry = Label(self.container_frame)
        self.save_file_entry.grid(column=1, row=5)
        self.save_file_entry.configure(text=workbook_file_name, bg='black', fg='orange', anchor='w')

        
        ### Buttons
        self.enter_data_button = Button(self.button_frame)
        self.enter_data_button.grid(column=0, row=0, padx=2)
        self.enter_data_button.configure(text='Save Data', fg='orange', bg=_gray)
        self.enter_data_button.configure(relief='raised', command=self.click_save)
    
        self.cancel_button = Button(self.button_frame)
        self.cancel_button.grid(column=2, row=0, padx=2)
        self.cancel_button.configure(text='Cancel', fg='orange', bg=_gray)
        self.cancel_button.configure(relief='raised', command=self.click_quit)
        
        
        
        self.root.mainloop()
        
    def click_quit(self):
        self.root.quit()
        self.root.destroy()
    def click_save(self, event=None):
        if (self.name_entry.get() != '' and self.date_entry.get() != '' and self.reps_entry.get() != '' and self.mass_entry.get() != '' and self.style_entry.get() != ''):
            save_x_data = []
            save_y_data = []
            save_mass = 0
            
            try:
                save_mass = int(self.mass_entry.get())
            except:
                print("Invalid mass entry") 
            
            try:
                save_reps = int(self.reps_entry.get())
            except:
                print("Invalid rep entry")
                
            if isinstance(save_mass, int):
                save_x_data.extend([self.name_entry.get(), self.date_entry.get(), save_mass, save_reps, self.style_entry.get(), 'x'])
                save_x_data.extend(self.tot_x_data)
                
                save_y_data.extend([self.name_entry.get(), self.date_entry.get(), save_mass, save_reps, self.style_entry.get(), 'y'])
                save_y_data.extend(self.tot_y_data)
                
                try:
                    wb = load_workbook(filename=workbook_file_name)
                except:
                    wb = Workbook()
                ws = wb.active
                cols = ws.max_column
                
                index_value = ws.cell(row=1, column=cols).value
                
                if index_value == 'Index':
                    index_value = 1
                else:
                    index_value = int(index_value) + 1
                
                save_x_data.insert(0, index_value)
                save_y_data.insert(0, index_value)
                
                for x in range(1, len(save_x_data)):
                    ws.cell(row=x, column=cols+1).value = save_x_data[x-1]
                    ws.cell(row=x, column=cols+2).value = save_y_data[x-1]
                    
                wb.save(workbook_file_name)
                
                self.root.quit()
                self.root.destroy()
        else:
            print("input error")
            

if __name__ == '__main__':
    GUI = MainGUI()