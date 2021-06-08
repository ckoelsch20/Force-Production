from datetime import datetime
import tkinter.font as tkFont
import tkinter as tk
from openpyxl import load_workbook
from openpyxl import Workbook

class Save_Data_GUI:
    def __init__(self, x, y, mass, style):
        light_gray = '#404040'
        bright_orange = '#ff8303'
        dark_orange = '#ff8303'
        slate = '#1b1a17'

        self.root = tk.Tk()
        self.root.geometry("350x300")
        self.root.title("Save Force Production Data")
        self.root.configure(bg=slate)

        self.workbook_file_name = 'data\\force_production_data_testing.xlsx'
        
        self.tot_x_data = x
        self.tot_y_data = y
        self.mass = mass
        self.lift_style = style
        
        title_message_font_style = tkFont.Font(size=20)
        date = datetime.today().strftime('%Y-%m-%d')
        
        ### FRAMES
        self.title_frame = tk.Frame(self.root)
        self.title_frame.grid(column=0, row=0)
        
        self.container_frame = tk.Frame(self.root)
        self.container_frame.grid(column=0, row=1)
        self.container_frame.configure(bg=slate)
        
        self.button_frame = tk.Frame(self.root)
        self.button_frame.grid(column=0, row=2)
        self.button_frame.configure(bg=slate, pady=10)
        
        ### MESSAGES
        self.title_message = tk.Message(self.title_frame, font=title_message_font_style)
        self.title_message.grid(column=0, row=0)
        self.title_message.configure(text="Save Data")
        self.title_message.configure(bg=slate, fg=bright_orange, width=300, padx=115)
        
        label_width = 7
        
        self.name_label = tk.Label(self.container_frame)
        self.name_label.grid(column=0, row=0)
        self.name_label.configure(text='Name: ', bg=slate, fg=bright_orange, width=label_width, anchor='w')
        
        self.mass_label = tk.Label(self.container_frame)
        self.mass_label.grid(column=0, row=1)
        self.mass_label.configure(text='Mass: ', bg=slate, fg=bright_orange, width=label_width, anchor='w')
        
        self.reps_label = tk.Label(self.container_frame)
        self.reps_label.grid(column=0, row=2)
        self.reps_label.configure(text='Reps: ', bg=slate, fg=bright_orange, width=label_width, anchor='w')
        
        self.lift_style_label = tk.Label(self.container_frame)
        self.lift_style_label.grid(column=0, row=3)
        self.lift_style_label.configure(text='Lift Style:', bg=slate, fg=bright_orange, width=label_width, anchor='w')
        
        self.date_label = tk.Label(self.container_frame)
        self.date_label.grid(column=0, row=4)
        self.date_label.configure(text='Date: ', bg=slate, fg=bright_orange, width=label_width, anchor='w')
        
        self.save_file_label = tk.Label(self.container_frame)
        self.save_file_label.grid(column=0, row=5)
        self.save_file_label.configure(text='File: ', bg=slate, fg=bright_orange, width=label_width, anchor='w')
        
        ### ENTRIES
        self.name_entry = tk.Entry(self.container_frame)
        self.name_entry.grid(column=1, row=0)
        self.name_entry.bind('<Return>', self.click_save)

        self.mass_entry = tk.Entry(self.container_frame)
        self.mass_entry.grid(column=1, row=1)
        self.mass_entry.insert(tk.END, self.mass)
        self.mass_entry.bind('<Return>', self.click_save)
        
        self.reps_entry = tk.Entry(self.container_frame)
        self.reps_entry.grid(column=1, row=2)
        self.reps_entry.bind('<Return>', self.click_save)

        self.style_entry = tk.Entry(self.container_frame)
        self.style_entry.grid(column=1, row=3)
        self.style_entry.insert(tk.END, self.lift_style)
        self.style_entry.bind('<Return>', self.click_save)
        
        self.date_entry = tk.Entry(self.container_frame)
        self.date_entry.grid(column=1, row=4)
        self.date_entry.insert(tk.END, date)
        self.date_entry.bind('<Return>', self.click_save)
        
        self.save_file_entry = tk.Label(self.container_frame)
        self.save_file_entry.grid(column=1, row=5)
        self.save_file_entry.configure(text=self.workbook_file_name, bg=slate, fg=bright_orange, anchor='w')

        ### BUTTONS
        self.enter_data_button = tk.Button(self.button_frame, text='Save Data', command=self.click_save)
        self.enter_data_button.grid(column=0, row=0, padx=2)
        self.enter_data_button.configure(fg=bright_orange, bg=slate, relief='raised', height=1, width=10)
    
        self.cancel_button = tk.Button(self.button_frame, text='Cancel', command=self.click_quit)
        self.cancel_button.grid(column=2, row=0, padx=2)
        self.cancel_button.configure(fg=bright_orange, bg=slate, relief='raised', height=1, width=10)
        
        self.root.mainloop()

    def click_save(self, event=None):
        if (self.name_entry.get() != '' and self.date_entry.get() != '' and self.reps_entry.get() != '' and self.mass_entry.get() != '' and self.style_entry.get() != ''):
            save_x_data = []
            save_y_data = []
            save_mass = 0
            
            try:
                save_mass = int(self.mass_entry.get())
            except:
                print("Invalid mass entry") 
            
            try:
                save_reps = int(self.reps_entry.get())
            except:
                print("Invalid rep entry")
                
            if isinstance(save_mass, int):
                save_x_data.extend([self.name_entry.get(), self.date_entry.get(), save_mass, save_reps, self.style_entry.get(), 'x'])
                save_x_data.extend(self.tot_x_data)
                
                save_y_data.extend([self.name_entry.get(), self.date_entry.get(), save_mass, save_reps, self.style_entry.get(), 'y'])
                save_y_data.extend(self.tot_y_data)
                
                try:
                    wb = load_workbook(self.workbook_file_name)
                except:
                    wb = Workbook()
                ws = wb.active
                cols = ws.max_column
                
                index_value = ws.cell(row=1, column=cols).value
                
                if index_value == None:
                    save_index_data = ['Index', 'Name', 'Date', 'Mass', 'Reps', 'Style', 'Data']
                    for i in range(0, len(save_index_data)):
                        ws.cell(row=i+1, column=cols).value = save_index_data[i]
                    index_value = 1
                elif index_value == 'Index':
                    index_value = 1
                else:
                    index_value = int(index_value) + 1
                
                save_x_data.insert(0, index_value)
                save_y_data.insert(0, index_value)

                for x in range(0, len(save_x_data)):
                    ws.cell(row=x+1, column=cols+1).value = save_x_data[x]
                    ws.cell(row=x+1, column=cols+2).value = save_y_data[x]
                    
                wb.save(self.workbook_file_name)
                
                self.root.quit()
                self.root.destroy()
        else:
            print("input error")


    def click_quit(self):
        self.root.quit()
        self.root.destroy()

if __name__ == "__main__":
    Save_Data_GUI([],[],199,'test')